### AUTOR: LUCAS VARGAS|| 22/03/2022 ###
from bs4 import BeautifulSoup
from os import path
import openpyxl, time
import os, os.path

### infos necessárias||needed info: data |placa | destino | (volume ou litros) | valor
### dhEmi | idUnidTransp | xMunFim | qCarga | vTPrest

#lista das tags procuradas || a list of needed tags
nInfo = ["dhEmi", "idUnidTransp", "xMunFim", "qCarga", "vTPrest"]
#cria uma lista vazia pra depois || creates an empty list for later usage
dados = []

#checa se chegou novo xml || checks if a new xml arrived
if path.exists("CTEN.xml"):
    #muda o nome do xml recebido pra ter controle || changes the name of received xml
    os.rename("CTEN.xml", "CTEA.xml")
    #le o xml || reads the XML file
    with open('CTEA.xml', 'r') as f:
        data = f.read()
    #abre o excel de destino || opens target excel
    wb = openpyxl.load_workbook("Planilha Teste.xlsx")
    sheet = wb.active

                                   
    #passa as info para o a biblioteca||parses xml info to the library
    Bs_data = BeautifulSoup(data, "xml")
     

    for info in nInfo: #loop que passa cada informação da var nInfo|| for loop for each var in nInfo 
        x = Bs_data.find(info) #acha todas as tags especificadas||finds all specified tags
        
        x = x.text #remove a tag do xml|| removes xml tag

        res = str(x)#transforma a var do bs em string || transforms bs var into string

        if info == "dhEmi":#se a informação for a data||if the info is the current date
            #split no char T que significa hora|| splits on char T that means time
            a = x.split("T")
            #fica só com a primeira parte || keeps only the first part
            res = a[0]
            
        print(res)#imprime pra motivos de debug || prints for debugging reasons
        dados.append(res)# coloca o dado na lista dados|| puts data into var dados
        
    print(dados)#for debugging

    r = sheet.max_row + 1
    c = 1
    for d in dados:#loop pra colocar as info no excel || for loops that outputs chosen data to excel
        print(d)
        cell = sheet.cell(row = r, column = c)
        cell.value = d
        c = c+1

    wb.save("Planilha Teste.xlsx")# salva || saves
    os.remove("CTEA.xml")#deleta xml atual pra nao dar bagunça || deletes current xml for organization    

